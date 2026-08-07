#!/usr/bin/env python
"""
Standalone script: import Nebula T-Shirt sign-up form responses as Shop Orders.

This is a completely standalone script — NOT a `manage.py` subcommand. It
bootstraps Django itself (same way manage.py does) so it can be run directly:

    python import_tshirt_orders.py /path/to/file.xlsx [options]

WHERE TO PUT THIS FILE
-----------------------
Drop it anywhere in the project (e.g. project root, next to manage.py, or in
a `scripts/` folder) — it does not need to live inside an app. It just needs
to be run from an environment where `apps` and `config` are importable
(i.e. the project root is on PYTHONPATH — running it from the project root
works out of the box, same as manage.py).

WHAT IT DOES
------------
Reads the "Nebula T-Shirt Sign Up (Responses)" spreadsheet and creates one
`Order` row per response, linked to a Product you specify (matched by name,
or auto-selected/created). It:

  - Maps "Shirt size" (S/M/L/XL) to the product's "Size" variant attribute,
    matching an existing variant by value (case-insensitive).
  - Maps "T-Shirt Back Text" to a custom field answer (default label
    "Back Name") on the order, IF the product has a matching custom field.
    Otherwise it's tacked onto the staff notes so the info isn't lost.
  - Maps "Wilaya" (free text, mixed case) to the wilaya code used by the
    WILAYA_CHOICES on the Order model (e.g. "Alger" -> "16").
  - Maps "Baladyia" -> baladiya, "Full Name" -> full_name,
    "Phone Number" -> phone.
  - Sets every imported order's status to "confirmed", regardless of what
    the trailing status column ("Column 7" in the raw sheet) says. The
    original form status (done / annuler / blank) is preserved in the
    order's staff notes for reference, it just no longer drives `status`.
  - Uses the Timestamp column as submitted_at (backfilled after creation,
    since submitted_at is auto_now_add — see NOTE below).
  - Skips rows that already look imported (same phone + same back text)
    so the script is safe to re-run.

USAGE
-----
    # Dry run first — prints what WOULD be created, writes nothing
    python import_tshirt_orders.py /path/to/Nebula_T-Shirt_Sign_Up__Responses_.xlsx --dry-run

    # Attach to a specific existing product by name (recommended)
    python import_tshirt_orders.py /path/to/file.xlsx --product "Nebula T-Shirt"

    # If no product name given, the script will:
    #   - use the only product whose name matches "t-shirt"/"tshirt"/
    #     "nebula"/"jersey" if exactly one is found
    #   - otherwise STOP and print the exact list of products that exist
    #     in the shop, so you can pass --product "<exact name>" — it will
    #     NOT silently create a placeholder product anymore (that caused
    #     orders to attach to a disconnected throwaway product last time)
    python import_tshirt_orders.py /path/to/file.xlsx

    # Only if you deliberately want a brand-new placeholder product created:
    python import_tshirt_orders.py /path/to/file.xlsx --create-placeholder-product

    # Point at a non-default settings module if needed
    DJANGO_SETTINGS_MODULE=config.settings python import_tshirt_orders.py /path/to/file.xlsx

NOTE ON submitted_at
---------------------
Order.submitted_at is `auto_now_add=True`, which means Django ignores
whatever value you pass in .objects.create() the *first* time — the field
always gets "now". To preserve the original spreadsheet timestamps, this
script creates the row first, then does a raw .update() on the queryset
(update() bypasses auto_now_add) to backfill the real timestamp.

REQUIREMENTS
------------
Run it with the same Python environment / virtualenv you run manage.py
with (it needs Django + your project's installed apps importable). It
reads the spreadsheet with pandas if available, otherwise falls back to
openpyxl directly — no extra installs needed either way.
"""

import argparse
import os
import sys
from datetime import datetime
from pathlib import Path


# ── Bootstrap Django (equivalent of what manage.py does) ───────────────────
def bootstrap_django():
    # Make sure the project root (this file's directory, or wherever
    # manage.py lives) is importable, same as manage.py's sys.path[0].
    project_root = Path(__file__).resolve().parent
    sys.path.insert(0, str(project_root))

    os.environ.setdefault("DJANGO_SETTINGS_MODULE", "config.settings")

    try:
        import django
    except ImportError as exc:
        raise SystemExit(
            "Couldn't import Django. Run this script with the same Python "
            "environment / virtualenv you use for manage.py (activate your "
            "venv first)."
        ) from exc

    django.setup()


bootstrap_django()

# These imports MUST happen after django.setup()
from django.utils import timezone  # noqa: E402
from apps.shop.models import Order, Product, WILAYA_CHOICES  # noqa: E402


# ── Wilaya text -> code lookup ──────────────────────────────────────────────
def build_wilaya_lookup():
    lookup = {}
    for code, label in WILAYA_CHOICES:
        lookup[label.strip().lower()] = code
    # A few common alternate spellings seen in real-world form answers
    aliases = {
        "alger": "16",
        "algiers": "16",
        "ain temouchent": "46",
        "aïn temouchent": "46",
        "medea": "26",
        "médéa": "26",
        "tizi ouzou": "15",
        "blida": "09",
    }
    lookup.update(aliases)
    return lookup


WILAYA_LOOKUP = build_wilaya_lookup()

# Every imported order is forced to "confirmed" per requirement — the
# original form status is kept in notes for reference only, not applied.
FORCED_STATUS = "confirmed"


# ── Spreadsheet reading ──────────────────────────────────────────────────────
def read_rows(path):
    """Read the sign-up sheet, return list of dict rows. Tries pandas, falls
    back to openpyxl directly so this works with either dependency set.
    If neither is installed, tries to auto-install openpyxl (the lighter
    of the two) before giving up with a clear instruction."""
    try:
        import pandas as pd
        df = pd.read_excel(path)
        df = df.rename(columns=lambda c: str(c).strip())
        return df.to_dict(orient="records")
    except ImportError:
        pass

    try:
        import openpyxl
    except ImportError:
        print("Neither pandas nor openpyxl is installed in this environment.")
        print("Attempting to install openpyxl automatically ...")
        import subprocess
        try:
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
            import openpyxl  # retry import after install
        except Exception:
            raise SystemExit(
                "Could not auto-install openpyxl. Please run this manually in "
                "the same environment/venv you use for manage.py, then re-run "
                "this script:\n\n"
                "    pip install openpyxl\n"
            )

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    headers = [str(c.value).strip() if c.value else "" for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        rows.append(dict(zip(headers, row)))
    return rows


def clean(val):
    if val is None:
        return ""
    s = str(val).strip()
    if s.lower() == "nan":
        return ""
    return s


# ── Product resolution ───────────────────────────────────────────────────────
def resolve_product(name_arg, allow_create):
    if name_arg:
        try:
            return Product.objects.get(name=name_arg)
        except Product.DoesNotExist:
            existing = "\n".join(f'  - "{p.name}"' for p in Product.objects.all())
            raise SystemExit(
                f'No product named "{name_arg}" found.\n\n'
                f"Products that DO exist in the shop:\n{existing or '  (none)'}\n\n"
                f"Pass one of the exact names above with --product."
            )

    candidates = (
        Product.objects.filter(name__icontains="t-shirt")
        | Product.objects.filter(name__icontains="tshirt")
        | Product.objects.filter(name__icontains="nebula")
        | Product.objects.filter(name__icontains="jersey")
    ).distinct()

    if candidates.count() == 1:
        print(f'Auto-matched product: "{candidates.first().name}" (id={candidates.first().id})')
        return candidates.first()

    existing = "\n".join(f'  - "{p.name}"' for p in Product.objects.all())

    if candidates.count() > 1:
        names = ", ".join(f'"{p.name}"' for p in candidates)
        raise SystemExit(
            f"Multiple possible products found ({names}).\n\n"
            f"Products that exist in the shop:\n{existing or '  (none)'}\n\n"
            f're-run with --product "<exact name>" to disambiguate.'
        )

    # Nothing matched at all — refuse to guess. Auto-creating a placeholder
    # silently produced a duplicate/disconnected product last time, so this
    # now requires an explicit opt-in flag.
    if not allow_create:
        raise SystemExit(
            "No matching product found by name, and no --product was given.\n\n"
            f"Products that exist in the shop:\n{existing or '  (none)'}\n\n"
            f"Re-run with --product \"<exact existing product name>\" to attach "
            f"these orders to a real product (recommended), or pass "
            f"--create-placeholder-product to intentionally create a new "
            f'placeholder called "Nebula T-Shirt" instead.'
        )

    print(
        'No matching product found — creating a placeholder Product '
        '"Nebula T-Shirt" with Size variants (S/M/L/XL, stock 0, stock '
        'tracking OFF), because --create-placeholder-product was passed. '
        'Edit it in the dashboard to add price, banner, description, etc.'
    )
    product = Product.objects.create(
        name="Nebula T-Shirt",
        description="Imported from the Nebula T-Shirt Sign-Up form.",
        price="0.00",
        category="jersey",
        track_stock=False,
        is_active=True,
        variant_config={
            "attributes": [{"name": "Size"}],
            "variants": [
                {"id": f"var_{i+1}", "attribute": "Size", "value": v, "stock": 0}
                for i, v in enumerate(["S", "M", "L", "XL"])
            ],
        },
        custom_fields=[
            {"label": "Back Name", "placeholder": "e.g. SMITH", "required": False}
        ],
    )
    return product


def match_size_value(product, size_raw):
    """Return the exact variant 'value' string the product uses for this
    size (so it matches whatever casing/format was configured), or the
    raw cleaned size if no variant config exists yet."""
    size_clean = clean(size_raw).upper()
    variants = (product.variant_config or {}).get("variants", [])
    for v in variants:
        if str(v.get("attribute", "")).strip().lower() == "size" \
           and str(v.get("value", "")).strip().upper() == size_clean:
            return v.get("value")
    return size_clean  # fall back to raw value; order still records it


def has_custom_field(product, label):
    for f in product.custom_fields or []:
        if f.get("label", "").strip().lower() == label.strip().lower():
            return True
    return False


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    parser = argparse.ArgumentParser(
        description="Import Nebula T-Shirt sign-up form responses as Shop Orders."
    )
    parser.add_argument("xlsx_path", type=str, help="Path to the sign-up .xlsx file")
    parser.add_argument(
        "--product", type=str, default="",
        help="Exact name of the Product these orders should attach to. "
             "If omitted, the script tries to find/create one called 'Nebula T-Shirt'.",
    )
    parser.add_argument(
        "--back-text-label", type=str, default="Back Name",
        help="Custom field label on the product used for the shirt's back text "
             "(default: 'Back Name'). If the product has no matching custom field, "
             "the back text is appended to the order's staff notes instead.",
    )
    parser.add_argument(
        "--dry-run", action="store_true",
        help="Print what would be imported without writing anything to the database.",
    )
    parser.add_argument(
        "--create-placeholder-product", action="store_true",
        help="If no existing product matches, intentionally create a new "
             '"Nebula T-Shirt" placeholder product instead of stopping. '
             "Off by default so orders never get silently attached to a "
             "throwaway product.",
    )
    args = parser.parse_args()

    if not Path(args.xlsx_path).exists():
        raise SystemExit(f"File not found: {args.xlsx_path}")

    rows = read_rows(args.xlsx_path)
    if not rows:
        print("No rows found in the spreadsheet.")
        return

    product = resolve_product(args.product, args.create_placeholder_product)
    use_custom_field = has_custom_field(product, args.back_text_label)

    created, skipped = 0, 0

    for i, row in enumerate(rows, start=1):
        full_name = clean(row.get("Full Name"))
        phone = clean(row.get("Phone Number"))
        wilaya_raw = clean(row.get("Wilaya"))
        baladiya = clean(row.get("Baladyia") or row.get("Baladiya"))
        back_text = clean(row.get("T-Shirt Back Text"))
        size_raw = clean(row.get("Shirt size"))
        status_raw = clean(row.get("Column 7")).lower()
        timestamp = row.get("Timestamp")

        if not full_name or not phone:
            print(f"Row {i}: missing name or phone, skipping.")
            skipped += 1
            continue

        # Idempotency check: same phone + same back text on this product = already imported
        already = False
        for o in Order.objects.filter(phone=phone, product=product):
            if use_custom_field:
                if o.custom_field_values.get(args.back_text_label, "") == back_text:
                    already = True
                    break
            else:
                if back_text and back_text in (o.notes or ""):
                    already = True
                    break
        if already:
            print(f"Row {i}: order for {full_name} ({phone}) already exists, skipping.")
            skipped += 1
            continue

        wilaya_code = WILAYA_LOOKUP.get(wilaya_raw.lower(), "")
        if not wilaya_code and wilaya_raw:
            print(f'Row {i}: could not map wilaya "{wilaya_raw}" to a known code — leaving blank.')

        size_value = match_size_value(product, size_raw)
        variant_values = {"Size": size_value} if size_value else {}

        status = FORCED_STATUS

        custom_field_values = {}
        notes_parts = []
        if use_custom_field and back_text:
            custom_field_values[args.back_text_label] = back_text
        elif back_text:
            notes_parts.append(f"Back text: {back_text}")
        if status_raw:
            notes_parts.append(f"Original form status: {status_raw}")
        notes = " | ".join(notes_parts)

        print(
            f'Row {i}: {full_name} | {size_value or "?"} | "{back_text}" | '
            f'{wilaya_raw} ({wilaya_code or "?"}) / {baladiya} | {phone} | status={status}'
        )

        if args.dry_run:
            created += 1
            continue

        order = Order.objects.create(
            product=product,
            product_name=product.name,
            variant_values=variant_values,
            quantity=1,
            custom_field_values=custom_field_values,
            full_name=full_name,
            email="",
            phone=phone,
            wilaya=wilaya_code,
            baladiya=baladiya,
            address="",
            coupon_code="",
            discount_amount=0,
            total_amount=0,
            status=status,
            notes=notes,
        )

        # Backfill the real submission timestamp (submitted_at is
        # auto_now_add, so .create() above always stamps "now" —
        # .update() bypasses that).
        if timestamp:
            try:
                if isinstance(timestamp, str):
                    ts = datetime.fromisoformat(timestamp)
                else:
                    ts = timestamp  # pandas.Timestamp / datetime
                if timezone.is_naive(ts):
                    ts = timezone.make_aware(ts, timezone.get_current_timezone())
                Order.objects.filter(pk=order.pk).update(submitted_at=ts)
            except Exception as e:
                print(f"Row {i}: could not parse timestamp ({timestamp!r}): {e}")

        created += 1

    verb = "Would create" if args.dry_run else "Created"
    print(f"\n{verb} {created} order(s), skipped {skipped} row(s). Product: {product.name} (id={product.id})")


if __name__ == "__main__":
    main()